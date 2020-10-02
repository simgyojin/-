from openpyxl import Workbook
import openpyxl
from datetime import date
from datetime import datetime
from datetime import time
import os

print('='*60)
print('근무관리 시스템입니다. 문의사항은 교진이한테 물어보세영~')
# 기존파일
work_file = openpyxl.load_workbook('근무관리.xlsx')

#make_excel_sheet(take_date()[1], work_file)
# 워크시트 얻는 함수
def make_excel_sheet(month,file):
    month_list=file.sheetnames
    this_month=str(month)+'월'
    if this_month not in month_list:
        work_file.create_sheet(this_month)
        work_file_sheet = work_file.get_sheet_by_name(this_month)
        work_file_sheet['A1']='날짜'
        work_file_sheet['B1']='출근시각'
        work_file_sheet['C1']='퇴근시각'
        work_file_sheet['D1']='총 근무 시간'
        work_file_sheet['E1']='비고'
        return work_file_sheet
    else:
        work_file_sheet = work_file.get_sheet_by_name('{}월'.format(month))
        return work_file_sheet

# 날짜 얻는 함수
def take_date():
    take_d = input('*입력하실 근무일의 년-월-일 을 입력해주세요!: ').split('-')
    return take_d



##################3

#워크시트파일
#wsfile = make_excel_sheet(take_date()[1],work_file)
global now_time
now = datetime.now()
now_time=now.strftime("%H:%M")




###################3
def take_want():
    take_w = input('*오늘 근무에 대한 입력이라면 [오늘] , 이전 근무에 대한 입력이라면 [이전] 을 입력해주세요!: ')
    if take_w == '오늘':
        datee = str(int(str(date.today())[-5:-3]))
        wsfile = make_excel_sheet(datee, work_file)
        row_len = wsfile.max_row
        
        
        want = input('*출근이라면 [출근], 퇴근이라면 [퇴근], 비고가 있다면 [비고]를 입력해주세요: ')
        if want == '출근':
            input_date(str(date.today()).split('-'), wsfile, row_len)
            input_start(wsfile, row_len)
            save_file()
            print('{} 출근 기록 되었습니다.'.format(now_time))
        elif want == '퇴근':
            input_end(wsfile, row_len)
            print('{}에 퇴근 기록 되었습니다.'.format(now_time))
            save_file()
        elif want == '비고':
            print('{}내용이 비고에 기록되었습니다.'.format(input_bigo(wsfile, row_len)))
            save_file()
        else:
            print('!!정확한 단어를 입력해주세요!!')
            take_want()
            
    ##############################
    elif take_w == '이전':
        while True:
            datee = take_date()
            wsfile = make_excel_sheet(int(datee[1]), work_file)
            row_len = wsfile.max_row
            
            start = input('{}월 {}일의 출근시간을 입력하세요![ex) 13:30]: '.format(datee[1],datee[2]))
            wsfile['A{}'.format(row_len+1)]='{}년 {}월 {}일'.format(datee[0], datee[1],datee[2])
            wsfile['B{}'.format(row_len+1)]=start
            save_file()
    
            end = input('{}월 {}일의 퇴근시간을 입력하세요! [ex) 17:50]: '.format(datee[1],datee[2]))
            wsfile['C{}'.format(row_len+1)]=end
            startt = datetime.strptime(start, '%H:%M')
            endd = datetime.strptime(end, '%H:%M')
            wsfile['D{}'.format(row_len+1)]=str(endd-startt)[:4]
            save_file()
            
            bigo = input('비고 내용을 입력하세요: ')
            wsfile['E{}'.format(row_len+1)]=bigo
            save_file()
            more = input('*더 입력하실 근무가 있다면 [계속] 아니라면 [종료]를 입력해주세요: ')
            if more == '종료':
                print('근무관리 시스템을 종료합니다. 감사합니다.')
                break
            elif more == '계속':
                continue
            else:
                print('!!정확한 단어를 입력해주세요!!')
                continue
            
    else:
        print('!!정확한 단어를 입력해주세요!!')
        take_want()


# 날짜 입력    
def input_date(date, file, row_len):
    file['A{}'.format(row_len+1)]='{}년 {}월 {}일'.format(date[0], date[1],date[2])

# 출근시각입력 wsfile
# input_start(wsfile)
def input_start(file,row_len):
    global now_time
    now_time=now.strftime("%H:%M")
    file['B{}'.format(row_len+1)]=now_time

# 퇴근시각입력
def input_end(file, row_len):
    endd = datetime.strptime(str(datetime.now())[11:16],'%H:%M')
    start_cell = file.cell(row=row_len,column=2).value
    startt = datetime.strptime(str(start_cell), '%H:%M')
    file['C{}'.format(row_len)]=str(datetime.now())[11:16]
    file['D{}'.format(row_len)]=str(endd-startt)[:4]

# 비고 입력
def input_bigo(file, row_len):
    bigo = input('*비고 내용을 입력하세요: ')
    file['E{}'.format(row_len)]=bigo
    return bigo


def save_file():
    work_file.save('근무관리.xlsx')
    
if __name__ == '__main__':
    take_want()
os.system('pause')


