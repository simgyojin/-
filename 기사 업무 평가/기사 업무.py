# -*- coding: utf-8 -*-
"""
Created on Mon May 25 16:50:46 2020

@author: SAMSUNG
"""
import os
import pandas as pd
from openpyxl import Workbook
import datetime as dt
#from datetime import date

while True:
    print('='*80)
    print('반갑습니다. 기사 배차 간격 통계 시스템입니다. 그만 두시려면 [종료]를 눌러주세요')
    file_name=input('▶궁금하신 파일 명을 입력하세요: ')
    if file_name == '종료':
        break
    qq = input('▶모든 시간을 통계 내겠습니까? (맞다면 [네] 아니라면 [아니오] 입력): ')
    print()
    if qq == '아니오':
        print('▶몇시부터 몇시까지의 시간을 통계내시겠습니까?')
        start_time=int(input('시작 시간(24시 표현법으로): '))
        end_time=int(input('종료 시간(24시 표현법으로): '))
        print()
        print('='*80)
        print()
        print('[{}]파일의 [{}시] 부터 [{}시] 까지의 배차간격을 알아보겠습니다.'
              .format(file_name, start_time, end_time))
    
    if qq == '네':
        start_time=int('00')
        end_time=int('24')
        print()
        print('='*80)
        print()
        print('[{}]파일의 [모든 시간] 동안의 배차간격을 알아보겠습니다.'
              .format(file_name, start_time, end_time))
        
    riders=pd.read_excel('{}.xlsx'.format(file_name))
    
    update_rider = []
    for name in list(riders['기사']):
        if str(name)!='nan':
            update_rider.append(name[name.rfind(']')+1:].strip())
    
    rider_list=list(set(update_rider))
    long_rider=update_rider
    start_list=list(riders['접수'])
    end_list=list(riders['배차'])
    
    all_bb=[]
    rider_dic=dict()
    for rider in rider_list:
        rr = rider.strip()
        rider_dic.update({rr:[]})


    for count in range(len(long_rider)):
        if str(end_list[count])!='nan' and str(start_list[count])!='nan':
            start= dt.datetime.strptime(str(start_list[count])[-8:], "%H:%M:%S")
            end= dt.datetime.strptime(str(end_list[count])+':00', "%H:%M:%S")
            if int(str(start)[10:13])>=start_time and int(str(start)[10:13])<= end_time:
                term = str(end-start)
                #term1= dt.datetime.strptime(term, "%H:%M:%S")
                if '-1' not in term:
                    real_term=int(term[2:4])
                    rider_dic['{}'.format(long_rider[count])].append(real_term)
                    all_bb.append(real_term)


    write_wb = Workbook()
    write_ws = write_wb.active
    write_ws['A1']='기사명'
    write_ws['B1']='평균 배차 간격'
    write_ws['C1']='평균 대비 배차간격(높을 수록 오래된 콜 잡은 것)'
    write_ws['D1']='콜 개수'
    write_ws['E1']='최대 배차 간격'
    write_ws['G1']='각각 모든 콜 배차간격'
    # 세로 column

    # 가로 row
    #for key, valuee in rider_dic.items():
    #   if len(valuee) == 0:
    #      del(rider_dic[key])

    all_mean=sum(all_bb)//len(start_list)
    columnz = 2
    for key, valuee in rider_dic.items():
        # 기사명    
        write_ws.cell(columnz,1, key)
        
        #평균 배차 간격
        if len(valuee) != 0:
            mean_bb=sum(valuee)//len(valuee)
            write_ws.cell(columnz,2, mean_bb)
            # 평균 대비 배차 간격
            mean_bcg=sum(valuee)/(len(start_list)-len(valuee))
            write_ws.cell(columnz,3, mean_bcg)
        
        else:
            write_ws.cell(columnz,2, 0)
            write_ws.cell(columnz,3, 0)
        
        # 최대 배차 간격
        if len(valuee) > 1:
            write_ws.cell(columnz,5,max(valuee))
        # 특정시간 콜 개수
        write_ws.cell(columnz,4, len(valuee))
        # 콜 배차간격
        for i in range(len(valuee)):
            write_ws.cell(columnz, i+7, valuee[i])
    
        columnz+=1
    
    print()
    print('='*80)
    print()
    print('※통계를 완료하였습니다.')
    save=input('▶통계 결과를 저장할 엑셀 파일 이름을 입력해 주세요: ')
    print()
    print('='*80)
    dirr = os.path.dirname(__file__)
    write_wb.save(os.path.join(dirr, '{}.xlsx'.format(save)))
    print()
    if qq == '아니오':
        print('★[{}]파일의 [{}시] 부터 [{}시] 까지 배차간격을 [{}]파일에 저장했습니다.★'
              .format(file_name, start_time, end_time, save))
        
    if qq == '네':
        print('★[{}]파일의 모든 배차간격을 [{}]파일에 저장했습니다.★'
              .format(file_name, save))
    last_input=input('※끝내시려면 [종료]를 계속하시려면 [계속]을 입력하세요.: ')
    print()
    if last_input=='종료':
        print('감사합니다. 기사 배차 통계 시스템을 종료합니다.')
        break

os.system('pause')
